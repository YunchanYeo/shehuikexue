"""Human-readable report from evaluation result."""

from __future__ import annotations

from pathlib import Path
from typing import Any

# 维度综合分（与需求文档一致）
DIMENSION_LABELS_ZH = {
    "fluency": "流利度",
    "richness": "丰富性",
    "vividness": "生动性",
    "total": "总分",
}

# 原始指标（与需求文档一致）
METRIC_LABELS_ZH = {
    "articulation_rate": "发音速率",
    "silent_pause_ratio": "无声停顿占比",
    "filled_pause_density": "填充词密度",
    "repetition_repair_density": "重复/修正密度",
    "type_token_ratio": "词汇多样性(TTR)",
    "mean_utterance_length": "平均句长",
    "connective_density": "连接词密度",
    "pitch_sd": "基频标准差",
    "intensity_sd": "强度标准差",
    "duration_sec": "录音时长(秒)",
}

# 批量 CSV 列顺序：(内部键, 表头中文名)
BATCH_CSV_COLUMNS: list[tuple[str, str]] = [
    ("index", "序号"),
    ("file_name", "文件名"),
    ("fluency", DIMENSION_LABELS_ZH["fluency"]),
    ("richness", DIMENSION_LABELS_ZH["richness"]),
    ("vividness", DIMENSION_LABELS_ZH["vividness"]),
    ("total", DIMENSION_LABELS_ZH["total"]),
    ("articulation_rate", METRIC_LABELS_ZH["articulation_rate"]),
    ("silent_pause_ratio", METRIC_LABELS_ZH["silent_pause_ratio"]),
    ("filled_pause_density", METRIC_LABELS_ZH["filled_pause_density"]),
    ("repetition_repair_density", METRIC_LABELS_ZH["repetition_repair_density"]),
    ("type_token_ratio", METRIC_LABELS_ZH["type_token_ratio"]),
    ("mean_utterance_length", METRIC_LABELS_ZH["mean_utterance_length"]),
    ("connective_density", METRIC_LABELS_ZH["connective_density"]),
    ("pitch_sd", METRIC_LABELS_ZH["pitch_sd"]),
    ("intensity_sd", METRIC_LABELS_ZH["intensity_sd"]),
    ("duration_sec", METRIC_LABELS_ZH["duration_sec"]),
    ("z_articulation_rate", "Z分·发音速率"),
    ("z_silent_pause_ratio", "Z分·无声停顿占比"),
    ("z_filled_pause_density", "Z分·填充词密度"),
    ("z_repetition_repair_density", "Z分·重复/修正密度"),
    ("z_type_token_ratio", "Z分·词汇多样性"),
    ("z_mean_utterance_length", "Z分·平均句长"),
    ("z_connective_density", "Z分·连接词密度"),
    ("z_pitch_sd", "Z分·基频标准差"),
    ("z_intensity_sd", "Z分·强度标准差"),
    ("transcript", "转写文本"),
    ("status", "状态"),
    ("error", "错误信息"),
]


def format_result(result: dict[str, Any], *, lang: str = "zh") -> str:
    m = result["raw_metrics"]
    s = result["scores"]
    transcript = result.get("transcript") or ""

    denoise_on = result.get("denoise", False)
    denoise_line = (
        f"  降噪: 已开启 (强度 {result.get('denoise_strength', 0.75):.0%})"
        if denoise_on
        else "  降噪: 未开启"
    )

    if lang == "zh":
        lines = [
            f"文件: {result['audio_path']}",
            f"转写: {transcript or '(空)'}",
            denoise_line,
            "",
            "—— 原始指标 ——",
            f"  发音速率 (音节/秒):     {m['articulation_rate']:.3f}",
            f"  无声停顿占比:           {m['silent_pause_ratio']:.3f}",
            f"  填充词密度:             {m['filled_pause_density']:.3f}",
            f"  重复/修正密度:          {m['repetition_repair_density']:.3f}",
            f"  词汇多样性 (TTR):       {m['type_token_ratio']:.3f}",
            f"  平均句长 (词/句):       {m['mean_utterance_length']:.3f}",
            f"  连接词密度:             {m['connective_density']:.3f}",
            f"  基频标准差 (Hz):        {m['pitch_sd']:.3f}",
            f"  强度标准差:             {m['intensity_sd']:.4f}",
            f"  时长 (秒):              {m['duration_sec']:.2f}",
            "",
            "—— 各指标 Z 分 (值−均值)/标准差 ——",
            *_format_metric_z_rows(result.get("z_scores") or {}),
            "",
            "—— 维度综合分（上列 Z 分加权，非再次标准化）——",
            f"  流利度:   {s['fluency']:.3f}  (=¼·(Z发音−Z停顿−Z填充−Z重复))",
            f"  丰富性:   {s['richness']:.3f}  (=⅓·(Z词汇多样+Z句长+Z连接词))",
            f"  生动性:   {s['vividness']:.3f}  (=½·(Z基频SD+Z强度SD))",
            f"  总分:     {s['total']:.3f}  (=三维算术平均)",
            "",
            _norms_footnote_zh(result.get("norms") or {}),
        ]
        return "\n".join(lines)

    lines = [
        f"파일: {result['audio_path']}",
        f"전사(중국어): {transcript or '(없음)'}",
        "",
        "—— 원시 지표 ——",
        f"  발음 속도 (음절/초):      {m['articulation_rate']:.3f}",
        f"  무성 휴지 비율:           {m['silent_pause_ratio']:.3f}",
        f"  충전어 밀도:              {m['filled_pause_density']:.3f}",
        f"  반복/수정 밀도:           {m['repetition_repair_density']:.3f}",
        f"  어휘 다양성 (TTR):        {m['type_token_ratio']:.3f}",
        f"  평균 문장 길이 (단어/문): {m['mean_utterance_length']:.3f}",
        f"  접속어 밀도:              {m['connective_density']:.3f}",
        f"  기본주파수 SD (Hz):       {m['pitch_sd']:.3f}",
        f"  강도 SD:                  {m['intensity_sd']:.4f}",
        f"  길이 (초):                {m['duration_sec']:.2f}",
        "",
        "—— 차원 점수 (Z 표준화) ——",
        f"  유창성:   {s['fluency']:.3f}",
        f"  풍부성:   {s['richness']:.3f}",
        f"  생동성:   {s['vividness']:.3f}",
        f"  총점:     {s['total']:.3f}",
    ]
    return "\n".join(lines)


_Z_LABELS_ZH = {
    key: METRIC_LABELS_ZH[key]
    for key in (
        "articulation_rate",
        "silent_pause_ratio",
        "filled_pause_density",
        "repetition_repair_density",
        "type_token_ratio",
        "mean_utterance_length",
        "connective_density",
        "pitch_sd",
        "intensity_sd",
    )
}


def _norms_footnote_zh(norms_info: dict[str, Any]) -> str:
    n = int(norms_info.get("cohort_n") or 0)
    path = norms_info.get("path") or "speech_eval/norms.yaml"
    if norms_info.get("calibrated"):
        return (
            f"注: Z 分使用常模文件中的总体均值/标准差（非本文件单独计算）。\n"
            f"    常模: {path}，标定样本数 N={n}。\n"
            f"    单条录音的标准差为 0，不能对单文件做 Z；请用 calibrate_norms.py 先标定。"
        )
    return (
        "注: 当前常模未用多样本标定（或为占位值），Z 分仅供调试。\n"
        "    请对文件夹运行: python calibrate_norms.py 你的录音文件夹/\n"
        "    单条音频无法得到总体标准差，不能单独生成有效 Z 常模。"
    )


def _format_metric_z_rows(z: dict[str, float]) -> list[str]:
    if not z:
        return ["  (无)"]
    rows = []
    for key, label in _Z_LABELS_ZH.items():
        if key in z:
            rows.append(f"  {label}: {z[key]:+.3f}")
    return rows


def format_filename_display(path: str, max_len: int = 48) -> str:
    name = path.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    if len(path) <= max_len:
        return path
    return f"…/{name}"


def _fmt4(value: float) -> str:
    return f"{value:.4f}"


def _fmt_z(value: float) -> str:
    return f"{value:+.4f}"


def batch_csv_headers() -> list[str]:
    return [label for _, label in BATCH_CSV_COLUMNS]


def result_to_batch_csv_row(
    result: dict[str, Any] | None,
    path: Path,
    *,
    index: int,
    error: str = "",
) -> dict[str, str]:
    """One row for batch CSV; keys are Chinese column headers."""
    labels = {key: label for key, label in BATCH_CSV_COLUMNS}
    empty = {label: "" for label in labels.values()}
    empty[labels["index"]] = str(index)
    empty[labels["file_name"]] = path.name
    empty[labels["status"]] = "失败"
    empty[labels["error"]] = error
    if result is None:
        return empty

    m = result["raw_metrics"]
    s = result["scores"]
    z = result.get("z_scores") or {}
    transcript = (result.get("transcript") or "").replace("\n", " ").strip()

    row = {
        labels["index"]: str(index),
        labels["file_name"]: path.name,
        labels["fluency"]: _fmt4(s["fluency"]),
        labels["richness"]: _fmt4(s["richness"]),
        labels["vividness"]: _fmt4(s["vividness"]),
        labels["total"]: _fmt4(s["total"]),
        labels["articulation_rate"]: _fmt4(m["articulation_rate"]),
        labels["silent_pause_ratio"]: _fmt4(m["silent_pause_ratio"]),
        labels["filled_pause_density"]: _fmt4(m["filled_pause_density"]),
        labels["repetition_repair_density"]: _fmt4(m["repetition_repair_density"]),
        labels["type_token_ratio"]: _fmt4(m["type_token_ratio"]),
        labels["mean_utterance_length"]: _fmt4(m["mean_utterance_length"]),
        labels["connective_density"]: _fmt4(m["connective_density"]),
        labels["pitch_sd"]: _fmt4(m["pitch_sd"]),
        labels["intensity_sd"]: _fmt4(m["intensity_sd"]),
        labels["duration_sec"]: f"{m['duration_sec']:.2f}",
        labels["z_articulation_rate"]: _fmt_z(z.get("articulation_rate", 0.0)),
        labels["z_silent_pause_ratio"]: _fmt_z(z.get("silent_pause_ratio", 0.0)),
        labels["z_filled_pause_density"]: _fmt_z(z.get("filled_pause_density", 0.0)),
        labels["z_repetition_repair_density"]: _fmt_z(
            z.get("repetition_repair_density", 0.0)
        ),
        labels["z_type_token_ratio"]: _fmt_z(z.get("type_token_ratio", 0.0)),
        labels["z_mean_utterance_length"]: _fmt_z(z.get("mean_utterance_length", 0.0)),
        labels["z_connective_density"]: _fmt_z(z.get("connective_density", 0.0)),
        labels["z_pitch_sd"]: _fmt_z(z.get("pitch_sd", 0.0)),
        labels["z_intensity_sd"]: _fmt_z(z.get("intensity_sd", 0.0)),
        labels["transcript"]: transcript,
        labels["status"]: "成功",
        labels["error"]: "",
    }
    return row


def format_batch_summary_report(
    rows: list[dict[str, str]],
    *,
    audio_dir: str,
    norms_path: str,
    cohort_n: int,
    ok_count: int,
    err_count: int,
) -> str:
    """Plain-text summary for batch evaluation (Chinese)."""
    headers = batch_csv_headers()
    name_col = headers[1]
    fluency_col = DIMENSION_LABELS_ZH["fluency"]
    richness_col = DIMENSION_LABELS_ZH["richness"]
    vividness_col = DIMENSION_LABELS_ZH["vividness"]
    total_col = DIMENSION_LABELS_ZH["total"]
    status_col = headers[-2]

    lines = [
        "══════════════════════════════════════════════════════",
        "              中文口语评估 · 批量结果汇总",
        "══════════════════════════════════════════════════════",
        "",
        f"录音文件夹: {audio_dir}",
        f"常模文件:   {norms_path}（标定样本 N={cohort_n}）",
        f"处理结果:   成功 {ok_count} 个，失败 {err_count} 个",
        "",
        "【计分公式】",
        "  流利度 = ¼×(Z发音速率 − Z无声停顿占比 − Z填充词密度 − Z重复/修正密度)",
        "  丰富性 = ⅓×(Z词汇多样性 + Z平均句长 + Z连接词密度)",
        "  生动性 = ½×(Z基频标准差 + Z强度标准差)",
        "  总分   = (流利度 + 丰富性 + 生动性) ÷ 3",
        "",
        "【维度分数一览】（按总分从高到低）",
        f"{'序号':>4}  {'文件名':<28} {fluency_col:>8} {richness_col:>8} {vividness_col:>8} {total_col:>8}",
        "─" * 72,
    ]

    ok_rows = [r for r in rows if r.get(status_col) == "成功"]
    ok_rows.sort(key=lambda r: float(r.get(total_col) or 0), reverse=True)

    for r in ok_rows:
        lines.append(
            f"{r[headers[0]]:>4}  {r[name_col]:<28} "
            f"{r[fluency_col]:>8} {r[richness_col]:>8} "
            f"{r[vividness_col]:>8} {r[total_col]:>8}"
        )

    failed = [r for r in rows if r.get(status_col) != "成功"]
    if failed:
        lines.extend(["", "【失败文件】"])
        for r in failed:
            lines.append(f"  · {r[name_col]}: {r.get(headers[-1], '')}")

    lines.extend(
        [
            "",
            "【列说明】详见同目录 评估汇总表.xlsx 或 评估汇总表.csv",
            "  · 发音速率: 去掉长无声段后，每秒音节数（越高越利索）",
            "  · 无声停顿占比: >0.2s 无声段占总时长比例（越高越卡顿）",
            "  · 填充词密度: 嗯/啊/就是/那个 等占比（越高越不流利）",
            "  · 重复/修正密度: 重复与自我修正占比（越高组织越困难）",
            "  · 词汇多样性(TTR): 不重复词/总词数（越高词汇越丰富）",
            "  · 平均句长: 每句平均词数（反映句子完整度）",
            "  · 连接词密度: 因为/所以/然后 等占比（越高越有条理）",
            "  · 基频标准差: 音高起伏（越大越生动）",
            "  · 强度标准差: 音量起伏（越大越有感染力）",
            "  · Z分列: (该值 − 群体均值) ÷ 群体标准差",
            "",
            "详细转写与各指标原始值请查看 结果_json/ 文件夹。",
            "══════════════════════════════════════════════════════",
        ]
    )
    return "\n".join(lines)


_TEXT_EXCEL_HEADERS = frozenset({"文件名", "转写文本", "状态", "错误信息"})


def _excel_cell_value(header: str, value: str):
    if value == "":
        return None
    if header == "序号":
        return int(value)
    if header in _TEXT_EXCEL_HEADERS:
        return value
    try:
        return float(str(value).replace("+", ""))
    except ValueError:
        return value


def write_batch_excel(
    rows: list[dict[str, str]],
    path: Path,
    *,
    audio_dir: str,
    norms_path: str,
    cohort_n: int,
    ok_count: int,
    err_count: int,
) -> None:
    """Write formatted .xlsx with Chinese headers (requires openpyxl)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    headers = batch_csv_headers()
    total_col = DIMENSION_LABELS_ZH["total"]
    status_col = headers[-2]

    wb = Workbook()
    ws = wb.active or wb.create_sheet("评估汇总")
    ws.title = "评估汇总"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ok_rows = [r for r in rows if r.get(status_col) == "成功"]
    ok_rows.sort(key=lambda r: float(r.get(total_col) or 0), reverse=True)
    failed_rows = [r for r in rows if r.get(status_col) != "成功"]
    ordered = ok_rows + failed_rows

    for row_idx, row in enumerate(ordered, start=2):
        for col_idx, header in enumerate(headers, start=1):
            raw = row.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=_excel_cell_value(header, raw))
            if header in (
                DIMENSION_LABELS_ZH["fluency"],
                DIMENSION_LABELS_ZH["richness"],
                DIMENSION_LABELS_ZH["vividness"],
                DIMENSION_LABELS_ZH["total"],
            ):
                cell.number_format = "0.0000"
            elif header.startswith("Z分·"):
                cell.number_format = "+0.0000"
            elif header not in _TEXT_EXCEL_HEADERS and header != "序号":
                cell.number_format = "0.0000"
            if header == "转写文本":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(ordered) + 1, 2)}"

    width_hints = {
        "序号": 6,
        "文件名": 28,
        "流利度": 10,
        "丰富性": 10,
        "生动性": 10,
        "总分": 10,
        "转写文本": 48,
        "状态": 8,
        "错误信息": 24,
    }
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width_hints.get(header, 14)

    info = wb.create_sheet("说明")
    info_lines = [
        ["中文口语评估 · 批量结果"],
        [f"录音文件夹: {audio_dir}"],
        [f"常模文件: {norms_path}（N={cohort_n}）"],
        [f"成功 {ok_count} 个，失败 {err_count} 个"],
        [],
        ["【计分公式】"],
        ["流利度", "= ¼×(Z发音速率 − Z无声停顿占比 − Z填充词密度 − Z重复/修正密度)"],
        ["丰富性", "= ⅓×(Z词汇多样性 + Z平均句长 + Z连接词密度)"],
        ["生动性", "= ½×(Z基频标准差 + Z强度标准差)"],
        ["总分", "= (流利度 + 丰富性 + 生动性) ÷ 3"],
        [],
        ["【指标说明】"],
        ["发音速率", "去掉长无声段后，每秒音节数（越高越利索）"],
        ["无声停顿占比", ">0.2s 无声段占总时长比例（越高越卡顿）"],
        ["填充词密度", "嗯/啊/就是/那个 等占比（越高越不流利）"],
        ["重复/修正密度", "重复与自我修正占比（越高组织越困难）"],
        ["词汇多样性(TTR)", "不重复词/总词数（越高词汇越丰富）"],
        ["平均句长", "每句平均词数（反映句子完整度）"],
        ["连接词密度", "因为/所以/然后 等占比（越高越有条理）"],
        ["基频标准差", "音高起伏（越大越生动）"],
        ["强度标准差", "音量起伏（越大越有感染力）"],
        ["Z分列", "(该值 − 群体均值) ÷ 群体标准差"],
    ]
    for r, line in enumerate(info_lines, start=1):
        for c, text in enumerate(line, start=1):
            cell = info.cell(row=r, column=c, value=text)
            if r == 1:
                cell.font = Font(bold=True, size=14)
            elif line and line[0].startswith("【"):
                cell.font = Font(bold=True)
    info.column_dimensions["A"].width = 18
    info.column_dimensions["B"].width = 64

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
